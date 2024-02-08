from openpyxl import load_workbook 
wb = load_workbook("C:\Excel\Attendence.xlsx")
ws = wb.active
date = int(input("Enter Date: "))
x = int(input("Classes Attended: "))
y = int(input("Total Classes: "))
c1=ws.cell(row=date, column=2) # type: ignore
c1.value = x 
c2=ws.cell(row=date, column=3) # type: ignore
c2.value = y  
wb.save("C:\Excel\Attendence.xlsx")
